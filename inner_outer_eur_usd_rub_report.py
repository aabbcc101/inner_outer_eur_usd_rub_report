# -*- coding: utf-8 -*-

# I wanna show how can I use python and so on here

#TAKS: show all payments I mean inner and outer in right order I mean split it.
# They can contains eur, usd and rub operations 
# As a result we can get a xlsx report with nice collors and borders automaticaly  

import copy
import psycopg2
from configparser import ConfigParser
from openpyxl import load_workbook
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side

def config(filename, section):

    parser = ConfigParser()
    parser.read(filename)

    db = {}

    if parser.has_section(section):
        params = parser.items(section)
        for param in params:
            db[param[0]] = param[1]
    else:
        raise Exception('Section {0} not found in the {1} file'.format(section, filename))

    return db
    
# Не заполненные поля оставляет пустыми, сортировка в нужной последовательности. good enough foer me    
def sort_list(correct_list, incorrect_list, which_the_call): 

    correct_copy = correct_list.copy()

    if len(correct_list) >= len(incorrect_list):
        #print('\n new list \n', len(correct_list),' ' , len(incorrect_list))   
  
        for correct_i in range(len(correct_list)):
            for incorrect_i in range(len(incorrect_list)):      
                if (correct_list[correct_i][0] == incorrect_list[incorrect_i][0]) and (incorrect_list[incorrect_i][1] == correct_list[correct_i][1]):     
                    correct_copy[correct_i].append(incorrect_list[incorrect_i][2])
                    #round() is here
                    correct_copy[correct_i].append(round(incorrect_list[incorrect_i][3], 2))                  
                    #print(correct_copy[correct_i])
                    break                   
    else:
        print('проверьте актуальность списка, сделайте запрос в ручную и найдите не достающую строку, которой        не хватает в correct_list')
 
    #checking_empty_place 
    for correct_copy_i in range(len(correct_copy)):
        if (len(correct_copy[correct_copy_i]) == 3 and which_the_call  == 1) or  (len(correct_copy[correct_copy_i]) == 5 and which_the_call  == 2) or  (len(correct_copy[correct_copy_i]) == 7 and which_the_call  == 3):
            correct_copy[correct_copy_i].append('-')
            correct_copy[correct_copy_i].append('-')

    return(correct_copy)    

    
def  create_select(year, first_month_is, until_time_select, currency_select):   
    the_select = ('select dh."type", db.option_name, count (id_digital), '
        '(coalesce(sum(amount), 0)) AS total_amount '
        'from information dh '
        'join tbinformation_operation_type_params db on db.operation_id = dh.operation_id '
        "where dh.my_date >= '" + year + "-" + first_month_is + "-01 00:00:00' and dh.my_date < '" + year + "-" + first_month_is  + "-01'" + until_time_select + " "
        "and action_id in ('-FULFILLED','-EXECUTED','-DONE') and currency in ('" + currency_select +"') "
        'group by dh."type",db.option_name;')
    return the_select    
    
def main():
    # default data
    year = '2022'
    first_month_is = '04'
    second_month_is = '05'
    third_month_is = '06'
    name_first_month_is = 'error'
    name_second_month_is = 'error'
    name_third_month_is = 'error'  

    # TESTING
    #test time or real data where is the end INTO SELECTS
    until_time_select = "::date  + interval '1 hour'"
    #until_time_select = "::date  + interval '1 week'"
    #until_time_select = "::date  + interval '1 month'"
    
    #  ---------------------   MENU  -----------------------
    print('\n\n  **************** Программа для формирования ежеквартального отчета  **************** \n\n')
    print('Не забудь закрыть new_report.xlsx иначе выдаст ошибку при записи в самом конце ожиданий!\n') 
    print(' ***************************************************************************************** \n')
    print("При первом запуске изменить Путь до паролей filename='C:\\Users\\vlad\\Documents\\DO_NOT_SHARE\\dbases.ini'") 
    print('файл с паролем содержит поля для парсинга:\n')
    print('[postgresql_database]')
    print('host=postgresql_database')
    print('port=5252')
    print('database=my_database')
    print('user=vlad')
    print('password=My_PASSWORD\n')  
    print('Для теста программы: раскоментировать until_time_select = "::date  + interval 1 hour" , закоментировать  ту, где 1 month\n\n')
    print(' ***************************************************************************************** \n')

    #-----------------------------------------------------
    
    correct_list = [
        ['PAYMENT_Template', 'PAYMENT_Template', 'внешний'],
        ['PAYMENT_2', 'PAYMENT_2', 'внешний'],
        ['PAYMENT_3', 'PAYMENT_3', 'внешний'],
        ['PAYMENT_10', 'PAYMENT_10', 'внешний'],
        ['PAYMENT_11', 'PAYMENT_11', 'внешний'],
        ['PAYMENT_16', 'PAYMENT_16', 'внешний'],
        ['PAYMENT_17', 'PAYMENT_17', 'внешний'],
        ['PAYMENT_20', 'PAYMENT_20', 'внешний'],
        ['PAYMENT_7', 'PAYMENT_7', 'внешний'],
        ['PAYMENT_31', 'PAYMENT_31', 'внешний'],
        ['PAYMENT_21', 'PAYMENT_21', 'внешний'],
        ['PAYMENT_32', 'PAYMENT_32', 'внешний'],
        ['PAYMENT_33', 'PAYMENT_33', 'внешний'],
        ['PAYMENT_34', 'PAYMENT_34', 'внешний'],
        ['PAYMENT_35', 'PAYMENT_35', 'внешний'],
        ['PAYMENT_36', 'PAYMENT_36', 'внешний'],
        ['PAYMENT_37', 'PAYMENT_37', 'внешний'],
        ['PAYMENT_38', 'PAYMENT_38', 'внешний'],
        ['PAYMENT_39', 'PAYMENT_39', 'внутренний'],
        ['PAYMENT_40', 'PAYMENT_40', 'внутренний'],
        ['PAYMENT_41', 'PAYMENT_42', 'внутренний'],
        ['PAYMENT_42', 'Открытие вклада', 'внутренний'],
        ['PAYMENT_43', 'PAYMENT_43', 'внутренний'],
        ['PAYMENT_44', 'PAYMENT_44', 'внутренний'],
        ['PAYMENT_8', 'PAYMENT_8', 'внутренний'],
        ['PAYMENT_9', 'PAYMENT_9', 'внутренний'],
        ['PAYMENT_18', 'PAYMENT_18', 'внутренний'],
        ['PAYMENT_19', 'PAYMENT_19', 'внутренний'],
        ['PAYMENT_51', 'PAYMENT_51', 'внутренний'],
        ['PAYMENT_52', 'PAYMENT_52', 'внутренний'],
        ['PAYMENT_53', 'PAYMENT_53', 'внутренний'],
        ['PAYMENT_54', 'PAYMENT_54', 'внутренний'],
        ['PAYMENT_55', 'PAYMENT_55', 'внутренний']
        ]
        
    # Параметры по умолчанию, чтобы не съезжали строки, когда изменяешь correct_list   
    
    #print(len(correct_list))
    length_of_columns  = 3 + len(correct_list)
    count_inner = 3
    for _, _, inner_or_external in correct_list:
        if inner_or_external == 'внешний':
            count_inner += 1
    
    # ENTERING YEAR
    
    year = -99
    while year == -99:
        try:
            year = int(input(' enter a year from 1980: '))
        except ValueError:
            year = -99
            print(f"You entered {year}, which is not a number")       
        if year > 1980 :
            year = str(year)
            pass            
        else:
            year = -99
            print('Введен не корректный год')

    # ENTERING QUARTER 
    
    months = [
        'Январь', 'Февраль','Март',
        'Апрель','Май','Июнь',
        'Июль','Август','Сентябрь',
        'Октябрь','Ноябрь','Декабрь'
        ]
        
    choose_a_month = -99
    while choose_a_month == -99:
        try:
            choose_a_month = int(input(' enter a month from 1 to 12: '))
        except ValueError:
            choose_a_month = -99
            print(f"You entered {choose_a_month}, which is not a number from 1 to 12")
    
        if choose_a_month in range(1, 13):
            first_month_is = '{0:02}'.format(choose_a_month)
            name_first_month_is = months[choose_a_month-1]
        else:
            choose_a_month = -99 
            print('Введен не корректный квартал')

    # DATABASE
    
    filename='C:\\Users\\vlad\\Documents\\DO_NOT_SHARE\\dbases.ini'
    section='postgresql_database'
    
    connect_postgresql = psycopg2.connect(**config(filename, section))
    print('connect_postgresql = ', connect_postgresql)
    cur = connect_postgresql.cursor()   
    
    #EXEL 
    wb_write = openpyxl.Workbook()
    wb_write.create_sheet(title = 'Первый лист', index = 0)
    sheet_write = wb_write['Первый лист']
    
    #width
    
    sheet_write.column_dimensions['A'].width = 16    
    sheet_write.column_dimensions['B'].width = 50
    #count - size
    for x in ('C', 'F', 'I'):
        sheet_write.column_dimensions[x].width = 10

    #surrency size
    for x in ('E', 'H', 'K'):
        sheet_write.column_dimensions[x].width = 9

    #total amount - size
    for x in ('G', 'D', 'J'):
        sheet_write.column_dimensions[x].width = 15
        
    #L
    sheet_write.column_dimensions['L'].width = 12 
    
    #color
    redFill = PatternFill(start_color='FFEE1111', end_color='FFEE1111', fill_type='solid')
    yellowFill = PatternFill(start_color='f6f495', end_color='f6f495', fill_type='solid')
    lawnGreen = PatternFill(start_color='bcfa61', end_color='bcfa61', fill_type='solid')
    mediumPurple = PatternFill(start_color='d5c2f9', end_color='d5c2f9', fill_type='solid')
    firstMonth = PatternFill(start_color='bffcc5', end_color='bffcc5', fill_type='solid')
    Rub_color = PatternFill(start_color='FFF8DC', end_color='FFF8DC', fill_type='solid')
    Usd_color = PatternFill(start_color='e0fdc4', end_color='e0fdc4', fill_type='solid')
    Eur_color = PatternFill(start_color='F0FFFF', end_color='F0FFFF', fill_type='solid')  
    
    
    for x in ('A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1', 'H1', 'I1', 'J1', 'K1', 'A2', 'B2', 'C2', 'D2', 'E2', 'F2', 'G2', 'H2', 'I2', 'J2', 'K2'): 
        sheet_write[x].fill = firstMonth
    
    for i in range(3, count_inner):
        sheet_write.cell(row=i, column=12).fill = lawnGreen   
    for i in range(count_inner, length_of_columns):
        sheet_write.cell(row=i, column=12).fill = mediumPurple  
        
    for i in range(3, length_of_columns):
        sheet_write.cell(row=i, column=5).fill = Rub_color  

    for i in range(3, length_of_columns):
        sheet_write.cell(row=i, column=8).fill = Usd_color  
          
    for i in range(3, length_of_columns):
        sheet_write.cell(row=i, column=11).fill = Eur_color  
        
        
    #sheet_write['E1'].fill = Rub_color
    
    #Borders
    thins = Side(border_style="thin", color="000000")
    #sheet_write['A1'].border = Border(top=thins, left=thins, bottom=thins, right=thins)
    sheet_write['A1'].border = Border(top=thins, left=thins, bottom=thins,)
    
    for x in ('B1', 'C1', 'D1', 'E1', 'F1', 'G1', 'H1', 'I1', 'J1'):
        sheet_write[x].border = Border(top=thins, bottom=thins)

    sheet_write['K1'].border = Border(right=thins)
    
    for i in range(2, length_of_columns):
        for j in range(1, 12):
            sheet_write.cell(row=i, column=j).border = Border(top=thins, left=thins, bottom=thins, right=thins) 
 
    for i in range(3, length_of_columns):
        sheet_write.cell(row=i, column=12).border = Border(top=thins, left=thins, bottom=thins, right=thins) 
    # values
    
    sheet_write['A1'] = name_first_month_is    
    sheet_write['A2'] = 'type'
    sheet_write['B2'] = 'option_name'
    
    for x in ('C2', 'F2', 'I2'):
        sheet_write[x] = 'count'
        
    for x in ('D2', 'G2', 'J2'):    
        sheet_write[x] = 'total_amount'

    for x in ('E2', 'H2', 'K2'):   
        sheet_write[x] = 'currency'
    
    E2_RUR = 'JPY'
    H2_USD = 'CNY'
    K2_EUR = 'KRW'

    currency_select1 = 'JPY'
    currency_select2 = 'CNY'
    currency_select3 = 'KRW'
 
    #SELECTS
    
    #month-1 rub
    the_select_1 = create_select(year, first_month_is, until_time_select, currency_select1)
    print(' loading first select... wait')    
    cur.execute(the_select_1)
    select_rows_1 = cur.fetchall()
    
    result = sort_list(correct_list, select_rows_1, 1)

    #month-1 USD
    the_select_1 = create_select(year, first_month_is, until_time_select, currency_select2)   
    print(' loading second select... wait a little more')     
    cur.execute(the_select_2)
    select_rows_2 = cur.fetchall()
    
    result = sort_list(result, select_rows_2, 2)
    
 
    #month-1 EUR
    the_select_1 = create_select(year, first_month_is, until_time_select, currency_select2)
    print(' loading third select... wait... almost ready')     
    cur.execute(the_select_3)
    select_rows_3 = cur.fetchall()
    
    result = sort_list(result, select_rows_3, 3)

    #Exel problems

    print('\n\n')
    print (len(result))
    print(result[0])
    print(result[0][1])
 
    for line in result:
        print(line)
    
    row_i = 3
    for name1, name2, inner_or_external, count_rub, amount_rub, count_usd, amount_usd, count_EUR, amount_EUR in result:
        print(name1, name2)
        cell = sheet_write.cell(row = row_i, column = 1)
        cell.value = str(name1)
        cell = sheet_write.cell(row = row_i, column = 2)
        cell.value = str(name2)
        cell = sheet_write.cell(row = row_i, column = 3)
        cell.value = str(count_rub)
        cell = sheet_write.cell(row = row_i, column = 4)
        cell.value = str(amount_rub)
        cell = sheet_write.cell(row = row_i, column = 5)
        cell.value = str(E2_RUR)
        cell = sheet_write.cell(row = row_i, column = 6)
        cell.value = str(count_usd)
        cell = sheet_write.cell(row = row_i, column = 7)
        cell.value = str(amount_usd)
        cell = sheet_write.cell(row = row_i, column = 8)
        cell.value = str(H2_USD)
        cell = sheet_write.cell(row = row_i, column = 9)
        cell.value = str(count_EUR)
        cell = sheet_write.cell(row = row_i, column = 10)
        cell.value = str(amount_EUR)
        cell = sheet_write.cell(row = row_i, column = 11)
        cell.value = str(K2_EUR)      
        cell = sheet_write.cell(row = row_i, column = 12)
        cell.value = str(inner_or_external)
        row_i += 1 
    
    wb_write.save('RESULT.xlsx')  
    #DATABASE Closing conection 
    cur.close()
    connect_postgresql.close()    

main()
